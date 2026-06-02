"""Unit tests for ExcelReportAdapter.

Tests verify:
- Locked 11-column set in Reconciliacion sheet (rev-3: +Año inferido)
- Summary sheet structure and per-registro aggregation
- CSV output column set
- Audit Trail sheet written only when non-empty
- ReportPort protocol conformance
- Output file is readable back via openpyxl (not just "no exception")
"""

from __future__ import annotations

import csv
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from reconciliation.adapters.report.xlsx_report import ExcelReportAdapter, _COLUMNS
from reconciliation.domain.models import ReconciliationRow
from reconciliation.domain.ports import ReportPort


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

LOCKED_COLUMNS = [
    "Registro",
    "Fecha",
    "Material",
    "Unidad",
    "Declarado",
    "Sumado(guías)",
    "Delta",
    "Estado",
    "Confianza mín",
    "Páginas origen",
    # Rev-3 D5 (REC-C07): advisory year-inference flag (EXT-021).
    "Año inferido",
]


def _make_row(
    registro: str = "232",
    fecha: date | None = date(2026, 5, 28),
    material: str = "barra a615/a706 g60 8mm (dob)",
    unidad: str = "TN",
    declared: str = "2.04",
    summed: str = "2.04",
    delta: str = "0",
    status: str = "MATCH",
    confidence: float | None = 0.99,
    pages: list[int] | None = None,
) -> ReconciliationRow:
    return ReconciliationRow(
        registro=registro,
        fecha=fecha,
        material_canonical=material,
        unidad=unidad,
        declared_qty=Decimal(declared),
        summed_qty=Decimal(summed),
        delta=Decimal(delta),
        status=status,  # type: ignore[arg-type]
        source_pages=pages or [3, 4],
        min_confidence=confidence,
    )


@pytest.fixture()
def sample_rows() -> list[ReconciliationRow]:
    return [
        _make_row(registro="232", material="barra a615/a706 g60 8mm (dob)", declared="2.04", summed="2.04", delta="0", status="MATCH"),
        _make_row(registro="232", material="barra a615/a706 g60 3/8\" (dob)", declared="7.163", summed="6.000", delta="1.163", status="MISMATCH", confidence=0.80),
        _make_row(registro="231", material="barra a615/a706 g60 8mm (dob)", declared="1.41", summed="1.41", delta="0", status="MATCH", confidence=None),
        _make_row(registro="231", material="barra a615/a706 g60 1\" (dob)", declared="2.966", summed="0", delta="2.966", status="DECLARED_MISSING", confidence=None),
    ]


@pytest.fixture()
def adapter() -> ExcelReportAdapter:
    return ExcelReportAdapter()


# ---------------------------------------------------------------------------
# Protocol conformance
# ---------------------------------------------------------------------------

class TestProtocolConformance:
    def test_implements_report_port(self, adapter: ExcelReportAdapter) -> None:
        assert isinstance(adapter, ReportPort)


# ---------------------------------------------------------------------------
# xlsx — column set
# ---------------------------------------------------------------------------

class TestXlsxColumnSet:
    def test_locked_columns_constant(self) -> None:
        """The _COLUMNS constant must match exactly the 11 locked columns (rev-3: +Año inferido)."""
        assert _COLUMNS == LOCKED_COLUMNS

    def test_reconciliacion_sheet_headers(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        dst = tmp_path / "report.xlsx"
        out = adapter.export(sample_rows, [], dst, "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        headers = [ws.cell(1, c).value for c in range(1, len(LOCKED_COLUMNS) + 1)]
        assert headers == LOCKED_COLUMNS

    def test_reconciliacion_row_count(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        dst = tmp_path / "report.xlsx"
        out = adapter.export(sample_rows, [], dst, "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        # 1 header + N data rows
        assert ws.max_row == 1 + len(sample_rows)

    def test_data_values_present(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        dst = tmp_path / "report.xlsx"
        out = adapter.export(sample_rows, [], dst, "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        # Row 2 (first data row) — check registro and material
        assert ws.cell(2, 1).value == "232"
        assert "barra" in (ws.cell(2, 3).value or "").lower()

    def test_fecha_iso_format(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        row = _make_row(fecha=date(2026, 5, 28))
        out = adapter.export([row], [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        assert ws.cell(2, 2).value == "2026-05-28"

    def test_fecha_none_becomes_empty_string(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        row = _make_row(fecha=None)
        out = adapter.export([row], [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        val = ws.cell(2, 2).value
        assert val is None or val == ""

    def test_confidence_none_becomes_empty(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        row = _make_row(confidence=None)
        out = adapter.export([row], [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        val = ws.cell(2, 9).value  # col 9 = Confianza mín
        assert val is None or val == ""

    def test_pages_comma_separated(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        row = _make_row(pages=[3, 10, 25])
        out = adapter.export([row], [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        val = ws.cell(2, 10).value  # col 10 = Páginas origen
        assert val is not None
        assert "3" in val and "10" in val and "25" in val


# ---------------------------------------------------------------------------
# xlsx — summary sheet
# ---------------------------------------------------------------------------

class TestXlsxSummarySheet:
    def test_resumen_sheet_exists(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        assert "Resumen" in wb.sheetnames

    def test_resumen_headers(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Resumen"]
        expected_headers = [
            "Registro", "Total líneas", "MATCH", "MISMATCH",
            "DECLARED_MISSING", "GUIA_MISSING", "UNCLASSIFIED",
        ]
        headers = [ws.cell(1, c).value for c in range(1, len(expected_headers) + 1)]
        assert headers == expected_headers

    def test_resumen_aggregates_correctly(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        """232 has 1 MATCH + 1 MISMATCH; 231 has 1 MATCH + 1 DECLARED_MISSING."""
        out = adapter.export(sample_rows, [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Resumen"]
        # Row 2 = registro "231" (sorted alphabetically first numerically: "231" < "232")
        row_231 = [ws.cell(2, c).value for c in range(1, 8)]
        row_232 = [ws.cell(3, c).value for c in range(1, 8)]
        # 231: total=2, MATCH=1, MISMATCH=0, DECLARED_MISSING=1
        assert row_231[0] == "231"
        assert row_231[1] == 2   # total
        assert row_231[2] == 1   # MATCH
        assert row_231[4] == 1   # DECLARED_MISSING
        # 232: total=2, MATCH=1, MISMATCH=1
        assert row_232[0] == "232"
        assert row_232[1] == 2
        assert row_232[2] == 1   # MATCH
        assert row_232[3] == 1   # MISMATCH

    def test_resumen_row_count(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Resumen"]
        # 2 distinct registros + 1 header
        assert ws.max_row == 3


# ---------------------------------------------------------------------------
# xlsx — audit trail sheet
# ---------------------------------------------------------------------------

class TestXlsxAuditTrail:
    def test_audit_trail_sheet_absent_when_empty(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        assert "Audit Trail" not in wb.sheetnames

    def test_audit_trail_sheet_present_when_non_empty(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        audit = [{"event": "edit", "registro": "232", "field": "declared_qty", "new_value": "2.5"}]
        out = adapter.export(sample_rows, audit, tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        assert "Audit Trail" in wb.sheetnames

    def test_audit_trail_headers_from_dict_keys(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        audit = [{"event": "edit", "field": "declared_qty"}]
        out = adapter.export(sample_rows, audit, tmp_path / "r.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Audit Trail"]
        assert ws.cell(1, 1).value == "event"
        assert ws.cell(1, 2).value == "field"


# ---------------------------------------------------------------------------
# xlsx — return value and file creation
# ---------------------------------------------------------------------------

class TestXlsxFileOutput:
    def test_export_returns_path(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.xlsx", "xlsx")
        assert isinstance(out, Path)

    def test_export_creates_xlsx_file(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.xlsx", "xlsx")
        assert out.exists()
        assert out.suffix == ".xlsx"

    def test_export_forces_xlsx_suffix(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        """Passing a .csv dst with fmt='xlsx' should still produce .xlsx."""
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "xlsx")
        assert out.suffix == ".xlsx"

    def test_empty_rows_produces_header_only_sheet(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        out = adapter.export([], [], tmp_path / "empty.xlsx", "xlsx")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Reconciliacion"]
        assert ws.max_row == 1  # header only


# ---------------------------------------------------------------------------
# csv — column set
# ---------------------------------------------------------------------------

class TestCsvOutput:
    def test_csv_creates_file(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "csv")
        assert out.exists()

    def test_csv_headers_match_locked_columns(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "csv")
        with out.open(encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            headers = next(reader)
        assert headers == LOCKED_COLUMNS

    def test_csv_row_count(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "csv")
        with out.open(encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh))
        # 1 header + N data rows
        assert len(rows) == 1 + len(sample_rows)

    def test_csv_resumen_file_created(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "csv")
        resumen_path = out.with_name("out_resumen.csv")
        assert resumen_path.exists()

    def test_csv_resumen_headers(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "csv")
        resumen_path = out.with_name("out_resumen.csv")
        with resumen_path.open(encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            headers = next(reader)
        expected = [
            "Registro", "Total líneas", "MATCH", "MISMATCH",
            "DECLARED_MISSING", "GUIA_MISSING", "UNCLASSIFIED",
        ]
        assert headers == expected

    def test_csv_data_values(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.csv", "csv")
        with out.open(encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            next(reader)  # skip header
            first_row = next(reader)
        assert first_row[0] == "232"
        assert "barra" in first_row[2].lower()

    def test_csv_forces_csv_suffix(
        self, adapter: ExcelReportAdapter, sample_rows: list[ReconciliationRow], tmp_path: Path
    ) -> None:
        out = adapter.export(sample_rows, [], tmp_path / "out.xlsx", "csv")
        assert out.suffix == ".csv"

    def test_csv_utf8_bom(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        """File should open as UTF-8-sig (Excel-compatible BOM)."""
        out = adapter.export([], [], tmp_path / "out.csv", "csv")
        raw = out.read_bytes()
        assert raw[:3] == b"\xef\xbb\xbf"  # UTF-8 BOM


# ---------------------------------------------------------------------------
# Error handling
# ---------------------------------------------------------------------------

class TestErrorHandling:
    def test_invalid_format_raises_value_error(
        self, adapter: ExcelReportAdapter, tmp_path: Path
    ) -> None:
        with pytest.raises(ValueError, match="Unsupported format"):
            adapter.export([], [], tmp_path / "out", "pdf")  # type: ignore[arg-type]

"""Integration-level E2E tests (tasks 6.1 + 6.2).

These tests exercise the full HTTP API round-trip using FastAPI TestClient
without running a real server, real OCR, or real vision LLM.  The pipeline
is bypassed via direct registry injection (same technique as test_api_routes.py).

6.1 - Happy path:
  POST /runs -> 202 + run_id
  GET /runs/{id} -> polls to review status
  GET /runs/{id}/table -> rows with 10-column DTO shape
  PATCH /runs/{id}/rows/{row_id} -> edit accepted, rows returned
  POST /runs/{id}/reassign -> reassignment accepted
  POST /runs/{id}/export -> xlsx download, 10 columns
  Cache-not-overwritten: second write on same ctx raises (abort/resume contract)

6.2 - Error paths:
  Corrupt PDF -> structured error, status=error surfaced in GET /runs/{id}
  UNCLASSIFIED kind surfaces in table rows
  OCR confidence < 0.85 -> requires_review=True in row
  Vision null date -> requires_review=True in row
  DECLARED_MISSING: guia without declared counterpart surfaces as DECLARED_MISSING
  GUIA_MISSING: declared without guia -> GUIA_MISSING
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from reconciliation.application.config import AppConfig
from reconciliation.domain.models import (
    GuiaContribution,
    GuiaDeRemision,
    MaterialLine,
    ReconciliationRow,
)
from reconciliation.infrastructure.api.main import create_app

# ---------------------------------------------------------------------------
# Shared helpers (mirrors test_api_routes.py pattern)
# ---------------------------------------------------------------------------


def _make_contribution(
    guia_id: str = "T001-0001",
    cantidad: str = "1000.0",
    unidad: str = "KG",
    confidence: float = 0.95,
) -> GuiaContribution:
    return GuiaContribution(
        guia_id=guia_id,
        source_pages=[5, 6],
        cantidad=Decimal(cantidad),
        unidad=unidad,
        confidence=confidence,
        identity_source="qr",
    )


def _make_row(
    registro: str = "232",
    material: str = "barra corrugada 1/2",
    unidad: str = "KG",
    declared: str = "1000.0",
    status: str = "MATCH",
    requires_review: bool = False,
) -> ReconciliationRow:
    d = Decimal(declared)
    contrib = _make_contribution(cantidad=declared)
    delta = Decimal("0") if status == "MATCH" else Decimal("100.0")
    return ReconciliationRow(
        registro=registro,
        fecha=date(2026, 5, 28),
        material_canonical=material,
        unidad=unidad,
        declared_qty=d,
        delta=delta,
        status=status,  # type: ignore[arg-type]
        source_pages=[5, 6],
        min_confidence=0.95,
        requires_review=requires_review,
        guias=[contrib],
    )


def _make_guia(
    guia_id: str = "T001-0001",
    registro: str | None = "232",
    fecha: date | None = date(2026, 5, 28),
) -> GuiaDeRemision:
    return GuiaDeRemision(
        guia_id=guia_id,
        registro=registro,
        fecha=fecha,
        fecha_confidence=0.95,
        lines=[
            MaterialLine(
                description_raw="BARRA CORRUGADA 1/2",
                description_canonical="barra corrugada 1/2",
                unidad="KG",
                cantidad=Decimal("1000.0"),
                confidence=0.95,
            )
        ],
        source_pages=[5, 6],
    )


def _make_review_service(
    rows: list[ReconciliationRow] | None = None,
    guias: list[GuiaDeRemision] | None = None,
    edit_result: list[ReconciliationRow] | None = None,
    reassign_result: list[ReconciliationRow] | None = None,
) -> MagicMock:
    default_rows = [_make_row()]
    svc = MagicMock()
    svc.rows = rows if rows is not None else default_rows
    svc.guias = guias if guias is not None else [_make_guia()]
    svc.get_audit_trail.return_value = []
    svc.apply_edit.return_value = edit_result if edit_result is not None else default_rows
    svc.apply_reassignment.return_value = (
        reassign_result if reassign_result is not None else default_rows
    )
    return svc


@pytest.fixture()
def app_client(tmp_path: Path) -> TestClient:
    """TestClient backed by a real FastAPI app with in-memory registry."""
    from reconciliation.infrastructure.run_history_store import (  # noqa: PLC0415
        JsonManifestRunHistoryAdapter,
    )

    app = create_app()
    config = AppConfig(output_dir=tmp_path / "runs")
    config.output_dir.mkdir(parents=True, exist_ok=True)
    app.state.config = config
    app.state.run_registry = {}
    # D1: this fixture bypasses the lifespan; seed the single run-history adapter
    # that _get_run_history resolves (normally created in main.lifespan).
    app.state.run_history = JsonManifestRunHistoryAdapter()
    return TestClient(app, raise_server_exceptions=True)


def _seed(
    client: TestClient,
    run_id: str,
    status: str = "review",
    svc: MagicMock | None = None,
    error: str | None = None,
    ctx: Any = None,
) -> None:
    """Inject a run entry into the registry (bypasses real pipeline)."""
    registry = client.app.state.run_registry  # type: ignore[attr-defined]
    registry[run_id] = {
        "status": status,
        "review_service": svc,
        "ctx": ctx,
        "result": None,
        "vision_calls_made": 2,
        "warnings": [],
        "error": error,
    }


# ---------------------------------------------------------------------------
# 6.1 - Happy-path E2E
# ---------------------------------------------------------------------------


class TestHappyPath:
    """Task 6.1: full API round-trip happy path."""

    def test_post_runs_returns_202_with_run_id(
        self, app_client: TestClient, tmp_path: Path
    ) -> None:
        """POST /runs with valid PDF bytes -> 202 Accepted, body has run_id + status."""
        fake_pdf = io.BytesIO(b"%PDF-1.4 1 0 obj<</Type/Catalog>>endobj")
        resp = app_client.post(
            "/api/v1/runs",
            files={"file": ("test.pdf", fake_pdf, "application/pdf")},
        )
        assert resp.status_code == 202, resp.text
        body = resp.json()
        assert "run_id" in body
        assert body["status"] in ("pending", "processing", "review", "error")

    def test_poll_status_review(self, app_client: TestClient) -> None:
        """GET /runs/{id} returns status=review after run completes."""
        run_id = str(uuid.uuid4())
        svc = _make_review_service()
        _seed(app_client, run_id, status="review", svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}")
        assert resp.status_code == 200
        body = resp.json()
        assert body["run_id"] == run_id
        assert body["status"] == "review"

    def test_get_table_returns_rows_with_10col_shape(self, app_client: TestClient) -> None:
        """GET /runs/{id}/table -> rows present, each row has the 10-column DTO shape."""
        run_id = str(uuid.uuid4())
        svc = _make_review_service()
        _seed(app_client, run_id, svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 200
        body = resp.json()
        assert body["run_id"] == run_id
        assert "rows" in body
        assert len(body["rows"]) >= 1

        row = body["rows"][0]
        # The 10 core columns of the DTO
        for field in (
            "row_id", "registro", "fecha", "material_canonical", "unidad",
            "declared_qty", "summed_qty", "delta", "status", "source_pages",
        ):
            assert field in row, f"Missing field: {field}"

    def test_patch_edit_row_returns_updated_rows(self, app_client: TestClient) -> None:
        """PATCH /runs/{id}/rows/{row_id} -> 200 with updated rows."""
        run_id = str(uuid.uuid4())
        updated_row = _make_row(status="MISMATCH")
        svc = _make_review_service(edit_result=[updated_row])
        _seed(app_client, run_id, svc=svc)

        # Use a simple row_id (the route param is not validated against actual rows)
        row_id = "r1"
        resp = app_client.patch(
            f"/api/v1/runs/{run_id}/rows/{row_id}",
            json={"guia_id": "T001-0001", "field": "fecha", "value": "2026-05-28"},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert "rows" in body

    def test_post_reassign_returns_updated_rows(self, app_client: TestClient) -> None:
        """POST /runs/{id}/reassign -> 200 with updated rows."""
        run_id = str(uuid.uuid4())
        svc = _make_review_service()
        _seed(app_client, run_id, svc=svc)

        resp = app_client.post(
            f"/api/v1/runs/{run_id}/reassign",
            json={
                "guia_id": "T001-0001",
                "new_registro": "231",
                "new_fecha": "2026-05-20",
            },
        )
        assert resp.status_code == 200
        body = resp.json()
        assert "rows" in body

    def test_post_export_xlsx_10_columns(self, app_client: TestClient, tmp_path: Path) -> None:
        """POST /runs/{id}/export -> xlsx download with exactly 10 header columns."""
        run_id = str(uuid.uuid4())
        svc = _make_review_service()
        fake_ctx = MagicMock()
        run_dir = tmp_path / "runs" / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        fake_ctx.run_dir = run_dir
        _seed(app_client, run_id, svc=svc, ctx=fake_ctx)

        # Build a minimal xlsx with the locked 10-column header
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Registro", "Fecha", "Material", "Unidad",
            "Declarado", "Sumado(guias)", "Delta", "Estado",
            "Confianza min", "Paginas origen",
        ]
        ws.append(headers)
        ws.append(["232", "2026-05-28", "barra", "KG", "1000", "1000", "0", "MATCH", "0.95", "5,6"])
        export_path = run_dir / "export.xlsx"
        wb.save(str(export_path))

        with patch("reconciliation.adapters.report.xlsx_report.ExcelReportAdapter") as MockExporter:
            instance = MockExporter.return_value
            instance.export.return_value = export_path
            resp = app_client.post(f"/api/v1/runs/{run_id}/export", json={"fmt": "xlsx"})

        assert resp.status_code == 200
        # Verify the returned xlsx content has 10 columns
        wb_check = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws_check = wb_check.active
        header_values = [
            ws_check.cell(1, col).value
            for col in range(1, ws_check.max_column + 1)
        ]
        header_values = [v for v in header_values if v is not None]
        assert len(header_values) == 10, (
            f"Expected 10 xlsx columns; got {len(header_values)}: {header_values}"
        )

    def test_extraction_cache_atomic_overwrite(self, tmp_path: Path) -> None:
        """Retry semantics: the cache is NOT write-once; a second write SUCCEEDS
        and fully replaces the content (SDD#3 PR-2). Atomicity (temp-file +
        rename) is preserved, so no partial/.tmp residue is left behind.

        This verifies the RunContext overwrite contract without running a full
        pipeline. The invariant is enforced at the RunContext level.
        """
        from reconciliation.application.run_context import RunContext  # noqa: PLC0415

        ctx = RunContext(pdf_path=tmp_path / "doc.pdf", output_base=tmp_path / "runs")
        ctx.write_extraction_cache({"run_id": ctx.run_id, "data": "first_write"})

        # Second write must SUCCEED and replace the content entirely.
        ctx.write_extraction_cache({"data": "second_write"})

        loaded = ctx.read_extraction_cache()
        assert loaded == {"data": "second_write"}
        # Atomicity preserved: no temp file left behind.
        assert list(ctx.run_dir.glob("*.tmp")) == []


# ---------------------------------------------------------------------------
# 6.2 - Error paths
# ---------------------------------------------------------------------------


class TestErrorPaths:
    """Task 6.2: structured error responses and flag propagation."""

    def test_error_run_surfaces_in_status(self, app_client: TestClient) -> None:
        """Corrupt PDF / pipeline error -> GET /runs/{id} returns status=error with detail."""
        run_id = str(uuid.uuid4())
        _seed(app_client, run_id, status="error", error="Corrupt PDF: no page count")

        resp = app_client.get(f"/api/v1/runs/{run_id}")
        assert resp.status_code == 200
        body = resp.json()
        assert body["status"] == "error"
        assert body["error"] is not None
        assert len(body["error"]) > 0

    def test_error_run_table_returns_422(self, app_client: TestClient) -> None:
        """GET /runs/{id}/table on an error run returns 422 (run not in review state)."""
        run_id = str(uuid.uuid4())
        _seed(app_client, run_id, status="error", error="Corrupt PDF")

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 422

    def test_unclassified_row_surfaces_in_table(self, app_client: TestClient) -> None:
        """UNCLASSIFIED page surfaces as row with status=UNCLASSIFIED in table (EXT-S04)."""
        run_id = str(uuid.uuid4())
        unclassified_row = _make_row(status="UNCLASSIFIED")
        svc = _make_review_service(rows=[unclassified_row])
        _seed(app_client, run_id, svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 200
        statuses = [r["status"] for r in resp.json()["rows"]]
        assert "UNCLASSIFIED" in statuses

    def test_low_confidence_row_has_requires_review(self, app_client: TestClient) -> None:
        """OCR confidence < 0.85 -> requires_review=True on the row (EXT-S08)."""
        run_id = str(uuid.uuid4())
        row_with_flag = _make_row(requires_review=True)
        svc = _make_review_service(rows=[row_with_flag])
        _seed(app_client, run_id, svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert "requires_review" in row
        assert row["requires_review"] is True

    def test_vision_null_date_row_has_requires_review(self, app_client: TestClient) -> None:
        """Vision null date -> requires_review=True on the row (EXT-S08b).

        ReconciliationService sets requires_review=True when guia.fecha is None.
        This test verifies the API surface propagates it correctly.
        """
        run_id = str(uuid.uuid4())
        row_null_date = _make_row(requires_review=True)
        svc = _make_review_service(rows=[row_null_date])
        _seed(app_client, run_id, svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert row["requires_review"] is True

    def test_declared_missing_surfaces_in_rows(self, app_client: TestClient) -> None:
        """Guia with no declared counterpart -> DECLARED_MISSING row in table (REC-S04)."""
        run_id = str(uuid.uuid4())
        dm_row = _make_row(status="DECLARED_MISSING")
        svc = _make_review_service(rows=[dm_row])
        _seed(app_client, run_id, svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 200
        statuses = [r["status"] for r in resp.json()["rows"]]
        assert "DECLARED_MISSING" in statuses

    def test_guia_missing_surfaces_in_rows(self, app_client: TestClient) -> None:
        """Declared with no guia -> GUIA_MISSING row in table (REC-S05)."""
        run_id = str(uuid.uuid4())
        gm_row = _make_row(status="GUIA_MISSING")
        svc = _make_review_service(rows=[gm_row])
        _seed(app_client, run_id, svc=svc)

        resp = app_client.get(f"/api/v1/runs/{run_id}/table")
        assert resp.status_code == 200
        statuses = [r["status"] for r in resp.json()["rows"]]
        assert "GUIA_MISSING" in statuses

    def test_unknown_run_returns_404(self, app_client: TestClient) -> None:
        """Any endpoint with unknown run_id returns 404."""
        fake_id = str(uuid.uuid4())
        resp = app_client.get(f"/api/v1/runs/{fake_id}")
        assert resp.status_code == 404

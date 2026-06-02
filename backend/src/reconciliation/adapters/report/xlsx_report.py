"""ExcelReportAdapter — ReportPort implementation writing xlsx and csv.

Locked 10-column set (EXT-003, per design/tasks):
    Registro | Fecha | Material | Unidad | Declarado | Sumado(guías) | Delta |
    Estado | Confianza mín | Páginas origen

xlsx output contains three sheets:
    1. Reconciliacion — the main 10-column reconciliation table
    2. Resumen         — per-registro summary (count, totals, status breakdown)
    3. Audit Trail     — the raw audit trail events (optional; omitted if empty)

csv output writes two UTF-8 files:
    <dst>.csv           — reconciliation table (same 10 columns)
    <dst>_resumen.csv   — summary sheet

Both xlsx and csv honour the same column order and data types.

openpyxl is used for xlsx (already declared in pyproject.toml).
stdlib csv for csv output.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from typing import Final, Literal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reconciliation.domain.models import ReconciliationRow

# ---------------------------------------------------------------------------
# Column specification (locked — EXT-003)
# ---------------------------------------------------------------------------

_COLUMNS: Final[list[str]] = [
    "Registro",
    "Fecha",
    "Material",
    "Unidad",
    "Declarado",
    "Sumado(guías)",
    "Delta",
    "Estado",
    "Confianza mín",
    "Páginas origen",
    # Rev-3 D5 (REC-C07): advisory year-inference flag (EXT-021).
    "Año inferido",
]

# Status → fill colour (ARGB)
_STATUS_FILLS: Final[dict[str, str]] = {
    "MATCH": "FF90EE90",          # light green
    "MISMATCH": "FFFFC0CB",       # light pink/red
    "DECLARED_MISSING": "FFFFA500",  # orange
    "GUIA_MISSING": "FFFFF44F",   # yellow
    "UNCLASSIFIED": "FFD3D3D3",   # light grey
}

_HEADER_FILL: Final[str] = "FF4F81BD"   # blue header
_HEADER_FONT_COLOUR: Final[str] = "FFFFFFFF"


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _row_to_values(row: ReconciliationRow) -> list[object]:
    """Serialise a ReconciliationRow to 11 ordered cell values (rev-3: +Año inferido)."""
    fecha_str = row.fecha.isoformat() if isinstance(row.fecha, date) else (row.fecha or "")
    pages_str = ", ".join(str(p) for p in sorted(row.source_pages)) if row.source_pages else ""
    conf_str = (
        f"{row.min_confidence:.2f}" if row.min_confidence is not None else ""
    )
    # Rev-3 D5: advisory year-inference flag (REC-C07).
    any_year_inferred_str = "Sí" if row.any_year_inferred else ""
    return [
        row.registro,
        fecha_str,
        row.material_canonical,
        row.unidad,
        str(row.declared_qty),
        str(row.summed_qty),
        str(row.delta),
        row.status,
        conf_str,
        pages_str,
        any_year_inferred_str,
    ]


def _build_summary(rows: list[ReconciliationRow]) -> list[list[object]]:
    """Build summary rows grouped by registro.

    Columns: Registro | Total líneas | MATCH | MISMATCH | DECLARED_MISSING |
             GUIA_MISSING | UNCLASSIFIED
    """
    from collections import defaultdict

    counts: dict[str, dict[str, int]] = defaultdict(
        lambda: {
            "total": 0,
            "MATCH": 0,
            "MISMATCH": 0,
            "DECLARED_MISSING": 0,
            "GUIA_MISSING": 0,
            "UNCLASSIFIED": 0,
        }
    )
    for row in rows:
        counts[row.registro]["total"] += 1
        counts[row.registro][row.status] += 1

    result: list[list[object]] = []
    for registro in sorted(counts):
        c = counts[registro]
        result.append([
            registro,
            c["total"],
            c["MATCH"],
            c["MISMATCH"],
            c["DECLARED_MISSING"],
            c["GUIA_MISSING"],
            c["UNCLASSIFIED"],
        ])
    return result


def _apply_header_style(ws: openpyxl.worksheet.worksheet.Worksheet, col_count: int) -> None:
    """Apply blue bold header style to row 1."""
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color=_HEADER_FONT_COLOUR)
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Set column widths based on the maximum content length."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)  # type: ignore[union-attr]
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ---------------------------------------------------------------------------
# xlsx writer
# ---------------------------------------------------------------------------


def _write_xlsx(
    rows: list[ReconciliationRow],
    audit_trail: list[dict],  # type: ignore[type-arg]
    dst: Path,
) -> Path:
    wb = openpyxl.Workbook()

    # --- Sheet 1: Reconciliacion ---
    ws_rec = wb.active
    ws_rec.title = "Reconciliacion"  # type: ignore[assignment]

    # Header row
    ws_rec.append(_COLUMNS)
    _apply_header_style(ws_rec, len(_COLUMNS))

    # Data rows
    for row in rows:
        values = _row_to_values(row)
        ws_rec.append(values)
        # Apply status fill to the whole row
        row_idx = ws_rec.max_row
        fill_colour = _STATUS_FILLS.get(row.status, "FFFFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_colour)
        for col_idx in range(1, len(_COLUMNS) + 1):
            ws_rec.cell(row=row_idx, column=col_idx).fill = row_fill

    _auto_width(ws_rec)

    # Freeze header row
    ws_rec.freeze_panes = "A2"

    # --- Sheet 2: Resumen ---
    ws_sum = wb.create_sheet("Resumen")
    summary_headers = [
        "Registro", "Total líneas", "MATCH", "MISMATCH",
        "DECLARED_MISSING", "GUIA_MISSING", "UNCLASSIFIED",
    ]
    ws_sum.append(summary_headers)
    _apply_header_style(ws_sum, len(summary_headers))
    for summary_row in _build_summary(rows):
        ws_sum.append(summary_row)
    _auto_width(ws_sum)
    ws_sum.freeze_panes = "A2"

    # --- Sheet 3: Audit Trail (omit if empty) ---
    if audit_trail:
        ws_audit = wb.create_sheet("Audit Trail")
        if audit_trail:
            audit_headers = list(audit_trail[0].keys())
            ws_audit.append(audit_headers)
            _apply_header_style(ws_audit, len(audit_headers))
            for event in audit_trail:
                ws_audit.append([str(v) for v in event.values()])
            _auto_width(ws_audit)

    # Ensure parent directory exists
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))
    return dst


# ---------------------------------------------------------------------------
# csv writer
# ---------------------------------------------------------------------------


def _write_csv(
    rows: list[ReconciliationRow],
    dst: Path,
) -> Path:
    """Write reconciliation table to *dst* (UTF-8 BOM for Excel compatibility)."""
    dst.parent.mkdir(parents=True, exist_ok=True)
    with dst.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(_COLUMNS)
        for row in rows:
            writer.writerow(_row_to_values(row))

    # Summary file
    summary_dst = dst.with_name(dst.stem + "_resumen.csv")
    summary_headers = [
        "Registro", "Total líneas", "MATCH", "MISMATCH",
        "DECLARED_MISSING", "GUIA_MISSING", "UNCLASSIFIED",
    ]
    with summary_dst.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(summary_headers)
        for summary_row in _build_summary(rows):
            writer.writerow(summary_row)

    return dst


# ---------------------------------------------------------------------------
# Public adapter
# ---------------------------------------------------------------------------


class ExcelReportAdapter:
    """Implements ReportPort — exports reconciliation results to xlsx or csv.

    For xlsx: writes Reconciliacion + Resumen + Audit Trail (if non-empty) sheets.
    For csv:  writes <dst>.csv + <dst>_resumen.csv (UTF-8 BOM).

    Both formats use the locked 10-column set defined in _COLUMNS.
    """

    def export(
        self,
        rows: list[ReconciliationRow],
        audit_trail: list[dict],  # type: ignore[type-arg]
        dst: Path,
        fmt: Literal["xlsx", "csv"],
    ) -> Path:
        """Export *rows* to *dst* in the specified format.

        Args:
            rows:        Reconciliation rows from ReconciliationService.
            audit_trail: Audit events from ReviewService (may be empty).
            dst:         Output path.  The suffix is ignored; the adapter
                         determines the correct extension.
            fmt:         "xlsx" or "csv".

        Returns:
            Path to the written file.

        Raises:
            ValueError: If *fmt* is not "xlsx" or "csv".
        """
        if fmt == "xlsx":
            xlsx_dst = dst.with_suffix(".xlsx")
            return _write_xlsx(rows, audit_trail, xlsx_dst)
        elif fmt == "csv":
            csv_dst = dst.with_suffix(".csv")
            return _write_csv(rows, csv_dst)
        else:
            raise ValueError(f"Unsupported format: {fmt!r}. Must be 'xlsx' or 'csv'.")
